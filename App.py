import docx
import openpyxl
import os
import tkinter as tk
from tkinter import filedialog, messagebox

def process_files(word_file_path, excel_file_path, output_folder="output"):
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):  
        data_dict = {}
        for idx, value in enumerate(row):
            column_header = sheet.cell(row=1, column=idx+1).value  
            data_dict[column_header] = value  

        new_doc = docx.Document(word_file_path)

        for paragraph in new_doc.paragraphs:
            for placeholder, value in data_dict.items():
                placeholder_tag = f"{{{placeholder}}}"
                if placeholder_tag in paragraph.text:
                    paragraph.text = paragraph.text.replace(placeholder_tag, str(value))

        # Get file name from the first column
        first_column_value = str(row[0]).strip() if row[0] else "Unnamed"
        name = first_column_value.replace(" ", "_")  # Remove spaces in filename

        output_word = os.path.join(output_folder, f"{name}.docx")
        
        counter = 1
        while os.path.exists(output_word):
            output_word = os.path.join(output_folder, f"{name}_{counter}.docx")
            counter += 1
        
        new_doc.save(output_word)
    
    messagebox.showinfo("Success", "Word files have been successfully generated.")

def browse_file(entry_widget, file_types):
    file_path = filedialog.askopenfilename(filetypes=file_types)
    entry_widget.delete(0, tk.END)
    entry_widget.insert(0, file_path)

def run_process():
    word_file_path = word_entry.get()
    excel_file_path = excel_entry.get()
    
    if not word_file_path or not excel_file_path:
        messagebox.showerror("Error", "Please select both Word and Excel files.")
        return
    
    try:
        process_files(word_file_path, excel_file_path)
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

root = tk.Tk()
root.title("Excel to Word Processor")
root.geometry("450x300")
root.configure(bg="#f0f0f0")

def create_labeled_entry(root, label_text, file_types, button_color):
    frame = tk.Frame(root, bg="#f0f0f0")
    frame.pack(pady=5, fill="x", padx=10)

    # Label
    label = tk.Label(frame, text=label_text, bg="#f0f0f0", font=("Arial", 10, "bold"))
    label.pack(anchor="w", padx=5)

    # Entry
    entry = tk.Entry(frame, width=50)
    entry.pack(pady=3)

    # Browse Button (Below Entry)
    button = tk.Button(frame, text="Browse", command=lambda: browse_file(entry, file_types), bg=button_color, fg="white", font=("Arial", 10, "bold"))
    button.pack(pady=3)

    return entry

excel_entry = create_labeled_entry(root, "Select Excel File:", [("Excel Files", ".xlsx;.xls")], "#217346")  # Green for Excel
word_entry = create_labeled_entry(root, "Select Word Template:", [("Word Files", ".docx;.doc")], "#2B579A")  # Blue for Word

process_button = tk.Button(root, text="Generate Word Files", command=run_process, bg="#FF8C00", fg="white", font=("Arial", 12, "bold"), pady=5)
process_button.pack(pady=20)

root.mainloop()
